"""
Módulo para operaciones de lectura y escritura de archivos Excel
"""
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


def cargar_hoja_excel(ruta_archivo, nombre_hoja):
    """
    Carga una hoja específica de un archivo Excel
    
    Args:
        ruta_archivo (str): Ruta al archivo Excel
        nombre_hoja (str): Nombre de la hoja a cargar
    
    Returns:
        pd.DataFrame: DataFrame con los datos o None si hay error
    """
    try:
        df = pd.read_excel(ruta_archivo, sheet_name=nombre_hoja)
        logger.info(f"Hoja '{nombre_hoja}' cargada correctamente del archivo: {ruta_archivo}")
        print(f"Hoja '{nombre_hoja}' cargada correctamente del archivo: {ruta_archivo}")
        
        if df.empty:
            mensaje = f"La hoja '{nombre_hoja}' del archivo Excel está vacía."
            logger.warning(mensaje)
            print(mensaje)
            return None
            
        return df
    except Exception as e:
        mensaje = f"Error al cargar la hoja '{nombre_hoja}' del archivo {ruta_archivo}: {str(e)}"
        logger.error(mensaje)
        print(mensaje)
        return None


def guardar_errores_consolidados(errores_por_hoja, ruta_salida):
    """
    Guarda los errores encontrados en un archivo Excel con múltiples hojas
    
    Args:
        errores_por_hoja (dict): Diccionario con hojas como claves y DataFrames de errores como valores
        ruta_salida (str): Ruta donde se guardará el archivo de errores
    """
    if not any(df is not None and not df.empty for df in errores_por_hoja.values()):
        logger.info(f"No hay errores para guardar en el archivo: {ruta_salida}")
        print(f"No hay errores para guardar en el archivo: {ruta_salida}")
        return
    
    try:
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            for hoja, df_errores in errores_por_hoja.items():
                if df_errores is not None and not df_errores.empty:
                    hoja_sin_punto = hoja.replace(".", "")
                    df_errores.to_excel(writer, sheet_name=hoja_sin_punto, index=False)
        
        logger.info(f"Archivo de errores consolidado guardado correctamente en: {ruta_salida}")
        print(f"Archivo de errores consolidado guardado correctamente en: {ruta_salida}")
    except Exception as e:
        mensaje = f"Error al guardar el archivo de errores consolidado: {str(e)}"
        logger.error(mensaje)
        print(mensaje)


def guardar_filas_con_multiples_errores(datos_originales, celdas_con_errores, ruta_salida):
    """
    Guarda las filas con múltiples errores en un archivo Excel
    
    Args:
        datos_originales (dict): Diccionario con hojas como claves y DataFrames originales como valores
        celdas_con_errores (dict): Diccionario con hojas como claves y diccionarios de celdas con errores
        ruta_salida (str): Ruta donde se guardará el archivo
    """
    try:
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            for hoja, df_original in datos_originales.items():
                if df_original is None or df_original.empty:
                    continue
                
                filas_con_multiples_errores = {}
                for fila_num, columnas in celdas_con_errores[hoja].items():
                    if len(columnas) >= 2:
                        filas_con_multiples_errores[fila_num] = columnas
                
                if not filas_con_multiples_errores:
                    continue
                
                indices_filas = [fila_num - 2 for fila_num in filas_con_multiples_errores.keys()]
                df_filtrado = df_original.iloc[indices_filas].copy()
                
                df_filtrado['Cantidad_Errores'] = [len(filas_con_multiples_errores[fila_num]) 
                                                  for fila_num in filas_con_multiples_errores.keys()]
                
                # guardar en Excel
                hoja_sin_punto = hoja.replace(".", "")
                df_filtrado.to_excel(writer, sheet_name=hoja_sin_punto, index=False)
                
                worksheet = writer.sheets[hoja_sin_punto]
                purple_fill = PatternFill(start_color='D8BFD8', end_color='D8BFD8', fill_type='solid')
                
                columnas_indices = {col: idx for idx, col in enumerate(df_filtrado.columns)}
                
                for i, (fila_num, columnas_error) in enumerate(filas_con_multiples_errores.items()):
                    for col_name in columnas_error:
                        if col_name in columnas_indices:
                            col_idx = columnas_indices[col_name]
                            cell = worksheet.cell(row=i+2, column=col_idx+1)
                            cell.fill = purple_fill
        
        logger.info(f"Archivo con filas de múltiples errores guardado correctamente en: {ruta_salida}")
        print(f"Archivo con filas de múltiples errores guardado correctamente en: {ruta_salida}")
    except Exception as e:
        mensaje = f"Error al guardar el archivo de filas con múltiples errores: {str(e)}"
        logger.error(mensaje)
        print(mensaje)


def guardar_datos_limpios(datos_originales, celdas_con_errores, ruta_salida):
    """
    Guarda un archivo Excel limpio sin las filas que contienen errores
    
    Args:
        datos_originales (dict): Diccionario con hojas como claves y DataFrames originales como valores
        celdas_con_errores (dict): Diccionario con hojas como claves y diccionarios de celdas con errores
        ruta_salida (str): Ruta donde se guardará el archivo limpio
    """
    try:
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            total_filas_eliminadas = 0
            
            for hoja, df_original in datos_originales.items():
                if df_original is None or df_original.empty:
                    continue
                
                # todas las filas con errores
                filas_con_errores = set(celdas_con_errores[hoja].keys())
                
                if not filas_con_errores:
                    # Si no hay errores, guardar el DataFrame completo
                    hoja_sin_punto = hoja.replace(".", "")
                    df_original.to_excel(writer, sheet_name=hoja_sin_punto, index=False)
                    logger.info(f"Hoja '{hoja}' guardada sin cambios (no se encontraron errores)")
                    print(f"Hoja '{hoja}' guardada sin cambios (no se encontraron errores)")
                else:
                    # fila_num es 2-based (fila 2 en Excel = índice 0 en DataFrame)
                    indices_sin_errores = []
                    for idx in range(len(df_original)):
                        fila_num = idx + 2  
                        if fila_num not in filas_con_errores:
                            indices_sin_errores.append(idx)
                    
                    df_limpio = df_original.iloc[indices_sin_errores].copy()
                    
                    hoja_sin_punto = hoja.replace(".", "")
                    df_limpio.to_excel(writer, sheet_name=hoja_sin_punto, index=False)
                    
                    filas_eliminadas = len(filas_con_errores)
                    total_filas_eliminadas += filas_eliminadas
                    
                    logger.info(f"Hoja '{hoja}': {filas_eliminadas} filas con errores eliminadas, "
                              f"{len(df_limpio)} filas guardadas")
                    print(f"Hoja '{hoja}': {filas_eliminadas} filas con errores eliminadas, "
                          f"{len(df_limpio)} filas guardadas")
        
        logger.info(f"Archivo de datos limpios guardado correctamente en: {ruta_salida}")
        logger.info(f"Total de filas eliminadas: {total_filas_eliminadas}")
        print(f"\nArchivo de datos limpios guardado correctamente en: {ruta_salida}")
        print(f"Total de filas eliminadas: {total_filas_eliminadas}")
        
    except Exception as e:
        mensaje = f"Error al guardar el archivo de datos limpios: {str(e)}"
        logger.error(mensaje)
        print(mensaje)